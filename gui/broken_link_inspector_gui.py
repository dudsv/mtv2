"""
Broken Link Inspector GUI component.

Checks for broken links on a single page or throughout a sitemap.
"""

import os
import datetime
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTextEdit, QRadioButton,
    QLineEdit, QLabel, QProgressBar, QFileDialog, QCheckBox, QMessageBox,
    QGroupBox
)
from openpyxl import Workbook

from workers.broken_link_worker import BrokenLinkWorker


class BrokenLinkInspectorGUI(QWidget):
    """Sub-tab 'Broken Link Inspector' within the Crawler."""
    
    def __init__(self):
        super().__init__()
        self.worker = None
        self.results = []
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout(self)

        title = QLabel("Broken Link Inspector")
        title.setObjectName("Title")
        layout.addWidget(title)

        # --------- Mode selection ----------
        mode_group = QGroupBox("Mode")
        mode_layout = QHBoxLayout()
        self.mode_single = QRadioButton("Single page checkup")
        self.mode_sitemap = QRadioButton("Sitemap audit (via sitemap.xml)")
        self.mode_single.setChecked(True)
        mode_layout.addWidget(self.mode_single)
        mode_layout.addWidget(self.mode_sitemap)
        mode_group.setLayout(mode_layout)
        layout.addWidget(mode_group)

        # --------- URL input ----------
        url_layout = QHBoxLayout()
        self.url_input = QLineEdit()
        self.url_input.setPlaceholderText("Ex (single): https://www.site.com/page")
        url_layout.addWidget(QLabel("URL:"))
        url_layout.addWidget(self.url_input)
        layout.addLayout(url_layout)

        # --------- Options ----------
        self.same_domain_cb = QCheckBox("Only same-domain links (single page)")
        self.same_domain_cb.setChecked(True)
        layout.addWidget(self.same_domain_cb)

        # --------- Controls ----------
        controls = QHBoxLayout()
        self.run_btn = QPushButton("Run check")
        self.run_btn.setProperty("accent", True)

        self.stop_btn = QPushButton("Stop")
        self.stop_btn.setEnabled(False)

        self.export_btn = QPushButton("Export Excel")
        self.export_btn.setEnabled(False)
        self.export_btn.setProperty("accent", False)

        self.clear_btn = QPushButton("Clear")

        controls.addWidget(self.run_btn)
        controls.addWidget(self.stop_btn)
        controls.addWidget(self.export_btn)
        controls.addWidget(self.clear_btn)
        layout.addLayout(controls)

        # --------- Progress + stats ----------
        self.progress = QProgressBar()
        layout.addWidget(self.progress)

        # Broken Link Inspector: link stats, not meta/H1
        self.stats_label = QLabel("Checked: 0 | OK: 0 | Redirect: 0 | 4xx: 0 | 5xx: 0 | Errors: 0")
        layout.addWidget(self.stats_label)

        # --------- Results ----------
        layout.addWidget(QLabel("Results (broken first):"))
        self.results_box = QTextEdit()
        self.results_box.setReadOnly(True)
        layout.addWidget(self.results_box)

        # --------- Log ----------
        layout.addWidget(QLabel("Log:"))
        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        layout.addWidget(self.log_box)

        # Connections
        self.mode_single.toggled.connect(self._on_mode_change)
        self.run_btn.clicked.connect(self.start_check)
        self.stop_btn.clicked.connect(self.stop_check)
        self.export_btn.clicked.connect(self.export_results)
        self.clear_btn.clicked.connect(self.clear_all)

        self._on_mode_change(self.mode_single.isChecked())

    def _on_mode_change(self, is_single: bool):
        if self.mode_single.isChecked():
            self.url_input.setPlaceholderText("Ex (single): https://www.site.com/page")
            self.same_domain_cb.setEnabled(True)
        else:
            self.url_input.setPlaceholderText("Ex (sitemap): https://www.site.com/sitemap.xml")
            self.same_domain_cb.setEnabled(False)

    def log(self, msg: str):
        self.log_box.append(msg)

    def set_export_ready(self, ready: bool):
        self.export_btn.setEnabled(ready)
        self.export_btn.setProperty("accent", ready)
        self.export_btn.style().unpolish(self.export_btn)
        self.export_btn.style().polish(self.export_btn)

    def clear_all(self):
        if self.worker and self.worker.isRunning():
            QMessageBox.warning(self, "Busy", "Stop the current run before clearing.")
            return
        self.results = []
        self.results_box.clear()
        self.log_box.clear()
        self.progress.setValue(0)
        self.stats_label.setText("Checked: 0 | OK: 0 | Redirect: 0 | 4xx: 0 | 5xx: 0 | Errors: 0")
        self.set_export_ready(False)

    def start_check(self):
        if self.worker and self.worker.isRunning():
            QMessageBox.warning(self, "Busy", "A check is already running.")
            return

        url = self.url_input.text().strip()
        if not url:
            QMessageBox.warning(self, "Input Error", "Please enter a URL.")
            return

        if not url.startswith("http://") and not url.startswith("https://"):
            url = "https://" + url

        mode = "single" if self.mode_single.isChecked() else "sitemap"
        same_domain = self.same_domain_cb.isChecked()

        self.results = []
        self.results_box.clear()
        self.progress.setValue(0)
        self.set_export_ready(False)

        self.log_box.clear()
        self.log(f"[START] Mode={mode}, URL={url}")

        self.run_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)

        self.worker = BrokenLinkWorker(mode=mode, root_url=url, same_domain_only=same_domain)
        self.worker.progress_update.connect(self.progress.setValue)
        self.worker.log_update.connect(self.log)
        self.worker.finished.connect(self.on_worker_finished)
        self.worker.start()

    def stop_check(self):
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.log("[WARN] Stop requested by user.")
            self.stop_btn.setEnabled(False)

    def on_worker_finished(self, results: list):
        self.results = results or []
        self.run_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)

        self._render_results()
        self.set_export_ready(bool(self.results))

    def _render_results(self):
        if not self.results:
            self.stats_label.setText("Checked: 0 | OK: 0 | Redirect: 0 | 4xx: 0 | 5xx: 0 | Errors: 0")
            self.results_box.setPlainText("No results.")
            return

        total = len(self.results)
        ok = sum(1 for r in self.results if r["category"] == "ok")
        redirect = sum(1 for r in self.results if r["category"] == "redirect")
        c4 = sum(1 for r in self.results if r["category"] == "client_error")
        c5 = sum(1 for r in self.results if r["category"] == "server_error")
        err = sum(1 for r in self.results if r["category"] == "network_error")

        self.stats_label.setText(
            f"Checked: {total} | OK: {ok} | Redirect: {redirect} | 4xx: {c4} | 5xx: {c5} | Errors: {err}"
        )

        def fmt(r):
            status = r["status"] if r["status"] is not None else "ERR"
            return f"[{status}] ({r['category']}) {r['url']}"

        broken_first = [
            r for r in self.results
            if r["category"] in ("client_error", "server_error", "network_error")
        ]
        redirects = [r for r in self.results if r["category"] == "redirect"]
        oks = [r for r in self.results if r["category"] == "ok"]

        lines = []

        if broken_first:
            lines.append("=== BROKEN / ERROR ===")
            lines.extend(fmt(r) for r in broken_first)
            lines.append("")

        if redirects:
            lines.append("=== REDIRECTS ===")
            lines.extend(fmt(r) for r in redirects)
            lines.append("")

        if oks:
            lines.append("=== OK ===")
            lines.extend(fmt(r) for r in oks)

        self.results_box.setPlainText("\n".join(lines))

    def export_results(self):
        if not self.results:
            QMessageBox.warning(self, "No data", "No results to export.")
            return

        folder = QFileDialog.getExistingDirectory(self, "Select folder to save broken link report")
        if not folder:
            self.log("[EXPORT] User cancelled export.")
            return

        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        mode = "single" if self.mode_single.isChecked() else "sitemap"
        filename = f"broken_links_{mode}_{ts}.xlsx"
        path = os.path.join(folder, filename)

        wb = Workbook()

        # Summary
        ws_sum = wb.active
        ws_sum.title = "Summary"
        ws_sum.append(["Metric", "Value"])

        total = len(self.results)
        ok = sum(1 for r in self.results if r["category"] == "ok")
        redirect = sum(1 for r in self.results if r["category"] == "redirect")
        c4 = sum(1 for r in self.results if r["category"] == "client_error")
        c5 = sum(1 for r in self.results if r["category"] == "server_error")
        err = sum(1 for r in self.results if r["category"] == "network_error")

        ws_sum.append(["Total checked", total])
        ws_sum.append(["OK (2xx)", ok])
        ws_sum.append(["Redirect (3xx)", redirect])
        ws_sum.append(["Client error (4xx)", c4])
        ws_sum.append(["Server error (5xx)", c5])
        ws_sum.append(["Network / other errors", err])

        # All results
        ws_all = wb.create_sheet("All")
        ws_all.append(["URL", "Status", "Category", "Final URL", "Error"])
        for r in self.results:
            ws_all.append([
                r["url"],
                r["status"],
                r["category"],
                r["final_url"],
                r["error"],
            ])

        # Broken only
        ws_broken = wb.create_sheet("Broken")
        ws_broken.append(["URL", "Status", "Category", "Final URL", "Error"])
        for r in self.results:
            if r["category"] in ("client_error", "server_error", "network_error"):
                ws_broken.append([
                    r["url"],
                    r["status"],
                    r["category"],
                    r["final_url"],
                    r["error"],
                ])

        wb.save(path)
        self.log(f"[EXPORT] Excel report saved to: {path}")
        QMessageBox.information(self, "Export", f"Report saved to:\n{path}")
