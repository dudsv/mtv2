"""
Sitemap Extractor GUI component.

Extracts URLs from sitemaps and allows comparison with external lists.
"""

import os
import asyncio
import aiohttp
import datetime
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTextEdit,
    QLineEdit, QLabel, QFileDialog, QMessageBox, QApplication
)
from openpyxl import Workbook

from config import HEADERS


class SitemapExtractorGUI(QWidget):
    """Sub-tab that extracts URLs from sitemaps and allows comparison with external list."""
    
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout(self)

        # ---------- STEP 1: SITEMAP ----------
        layout.addWidget(QLabel("Step 1 – Sitemap Index URL:"))

        top_row = QHBoxLayout()

        self.input_url = QLineEdit()
        self.input_url.setPlaceholderText("Ex: https://www.site.com/sitemap.xml")

        # Main purple button
        self.load_btn = QPushButton("Load Sitemaps")
        self.load_btn.setProperty("accent", True)

        self.clear_btn = QPushButton("Clear")

        top_row.addWidget(self.input_url)
        top_row.addWidget(self.load_btn)
        top_row.addWidget(self.clear_btn)
        layout.addLayout(top_row)

        # ---------- LOG OUTPUT ----------
        layout.addWidget(QLabel("Log Output:"))
        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        layout.addWidget(self.log_box)

        # ---------- STATS + EXTRACTED URLS ----------
        self.stats_label = QLabel("Pages: 0 | Title OK: 0 | Desc OK: 0 | OG Title OK: 0 | OG Desc OK: 0 | H1 OK: 0")
        layout.addWidget(self.stats_label)

        layout.addWidget(QLabel("Step 2 – Extracted URLs (from sitemap):"))
        self.result_box = QTextEdit()
        self.result_box.setReadOnly(False)
        layout.addWidget(self.result_box)

        export_layout = QHBoxLayout()
        self.btn_tsv = QPushButton("Copy TSV")
        self.btn_csv = QPushButton("Copy CSV")
        self.btn_list = QPushButton("Copy List")
        export_layout.addWidget(self.btn_tsv)
        export_layout.addWidget(self.btn_csv)
        export_layout.addWidget(self.btn_list)
        layout.addLayout(export_layout)

        # ---------- COMPARE SECTION ----------
        layout.addWidget(QLabel("Step 3 – Compare with Excel URLs:"))
        self.compare_box = QTextEdit()
        self.compare_box.setPlaceholderText("Cole aqui as URLs vindas do Excel, uma por linha")
        layout.addWidget(self.compare_box)

        self.compare_btn = QPushButton("Compare")
        self.compare_btn.setEnabled(False)              # Starts disabled
        self.compare_btn.setProperty("accent", False)   # No purple yet
        layout.addWidget(self.compare_btn)

        # ---------- DETAILED COMPARE OUTPUT ----------
        layout.addWidget(QLabel("Compare output:"))
        self.compare_output_box = QTextEdit()
        self.compare_output_box.setReadOnly(True)
        layout.addWidget(self.compare_output_box)

        # ---------- CONNECTIONS ----------
        self.load_btn.clicked.connect(self.run_extractor)
        self.clear_btn.clicked.connect(self.clear_all)

        self.btn_list.clicked.connect(self.copy_list)
        self.btn_tsv.clicked.connect(self.copy_tsv)
        self.btn_csv.clicked.connect(self.copy_csv)
        self.compare_btn.clicked.connect(self.compare_lists)

    def set_compare_ready(self, ready: bool):
        """Enable/disable the Compare button and apply purple styling when ready."""
        self.compare_btn.setEnabled(ready)
        self.compare_btn.setProperty("accent", ready)
        # Force Qt to reapply the style when the property changes
        self.compare_btn.style().unpolish(self.compare_btn)
        self.compare_btn.style().polish(self.compare_btn)

    # ---------- Log helpers ----------
    def log(self, txt):
        self.log_box.append(txt)

    # ---------- Main flow ----------
    def clear_all(self):
        self.result_box.clear()
        self.log_box.clear()
        self.compare_box.clear()
        self.compare_output_box.clear()
        self.stats_label.setText("Sub-sitemaps: 0 | URLs: 0 | Unique: 0")
        self.set_compare_ready(False)

    async def fetch_xml(self, url, session):
        try:
            async with session.get(url, ssl=False) as r:
                if r.status != 200:
                    self.log(f"[ERROR] {url} – status {r.status}")
                    return None
                return await r.text()
        except Exception as e:
            self.log(f"[ERROR] {url} – {e}")
            return None

    async def run_async(self, url):
        import xml.etree.ElementTree as ET
        urls, submaps = [], []

        async with aiohttp.ClientSession(headers=HEADERS) as session:
            self.log("[INIT] Fetching sitemap root...")
            xml_root = await self.fetch_xml(url, session)
            if not xml_root:
                return [], []

            try:
                root = ET.fromstring(xml_root)
            except Exception as e:
                self.log(f"[ERROR] Could not parse root sitemap XML: {e}")
                return [], []

            # Remove namespace from tag if it exists
            root_tag = root.tag.split('}')[-1].lower()

            # ---------- CASE 1: sitemapindex (has sub-sitemaps) ----------
            if root_tag == "sitemapindex":
                self.log("[INFO] Root is <sitemapindex> (has sub-sitemaps).")

                # Get only locs inside <sitemap>
                for sm_loc in root.iterfind(".//{*}sitemap/{*}loc"):
                    if sm_loc.text:
                        sm_url = sm_loc.text.strip()
                        submaps.append(sm_url)

                self.log(f"[INFO] Found {len(submaps)} sub-sitemaps.")

                for i, sm in enumerate(submaps, 1):
                    self.log(f"[FETCH] ({i}/{len(submaps)}) {sm}")
                    xml_sub = await self.fetch_xml(sm, session)
                    if not xml_sub:
                        continue

                    try:
                        r = ET.fromstring(xml_sub)
                        # Here we expect a <urlset> with <url><loc>...</loc></url>
                        for loc in r.iterfind(".//{*}url/{*}loc"):
                            if loc.text:
                                urls.append(loc.text.strip())
                    except Exception as e:
                        self.log(f"[PARSE ERROR] Could not parse sub-sitemap {sm}: {e}")

            # ---------- CASE 2: urlset (flat sitemap, no sub-sitemaps) ----------
            elif root_tag == "urlset":
                self.log("[INFO] Root is <urlset> (no nested sub-sitemaps).")
                for loc in root.iterfind(".//{*}url/{*}loc"):
                    if loc.text:
                        urls.append(loc.text.strip())

            # ---------- CASE 3: other format (generic fallback) ----------
            else:
                self.log(f"[WARN] Unknown root tag '{root_tag}', using generic <loc> parsing.")
                for loc in root.iterfind(".//{*}loc"):
                    if loc.text:
                        urls.append(loc.text.strip())

        return urls, submaps

    def run_extractor(self):
        url = self.input_url.text().strip()
        if not url:
            QMessageBox.warning(self, "Error", "Enter a sitemap URL.")
            return

        # Partial reset
        self.result_box.clear()
        self.compare_output_box.clear()
        self.stats_label.setText("Sub-sitemaps: 0 | URLs: 0 | Unique: 0")
        self.set_compare_ready(False)

        self.log_box.clear()
        self.log(f"[START] Processing sitemap: {url}")

        # We're in the UI thread => use asyncio.run
        try:
            asyncio.run(self._exec(url))
        except RuntimeError:
            # Fallback in case some loop already exists
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(self._exec(url))
            loop.close()

    async def _exec(self, url):
        urls, submaps = await self.run_async(url)
        unique = sorted(set(urls))
        self.result_box.setPlainText("\n".join(unique))
        self.stats_label.setText(
            f"Sub-sitemaps: {len(submaps)} | URLs: {len(urls)} | Unique: {len(unique)}"
        )
        self.set_compare_ready(bool(unique))
        self.log("[DONE] Sitemap extraction complete.")

    # ---------- Export helpers ----------
    def copy_list(self):
        QApplication.clipboard().setText(self.result_box.toPlainText())

    def copy_tsv(self):
        lines = [l for l in self.result_box.toPlainText().splitlines() if l.strip()]
        QApplication.clipboard().setText("\n".join(f"{l}\t" for l in lines))

    def copy_csv(self):
        lines = [l for l in self.result_box.toPlainText().splitlines() if l.strip()]
        QApplication.clipboard().setText("\n".join(f"{l}," for l in lines))

    # ---------- Compare flow ----------
    def compare_lists(self):
        extracted_raw = self.result_box.toPlainText().strip()
        if not extracted_raw:
            QMessageBox.warning(self, "No data", "Load a sitemap first to have URLs to compare.")
            return

        extracted = set(l.strip() for l in extracted_raw.splitlines() if l.strip())
        pasted = [v.strip() for v in self.compare_box.toPlainText().splitlines() if v.strip()]

        if not pasted:
            QMessageBox.warning(self, "No input", "Paste at least one URL from Excel to compare.")
            return

        found, missing = [], []
        for p in pasted:
            if p in extracted:
                found.append(p)
            else:
                missing.append(p)

        # ---- Detailed report for the user (in their own field) ----
        report_lines = []
        report_lines.append("===== COMPARE REPORT =====")
        report_lines.append(f"Total pasted (Excel): {len(pasted)}")
        report_lines.append(f"Found in sitemap:     {len(found)}")
        report_lines.append(f"Not found (missing):  {len(missing)}")
        report_lines.append("")

        report_lines.append("--- FOUND (present in sitemap) ---")
        if found:
            report_lines.extend(f"✔ {u}" for u in found)
        else:
            report_lines.append("(none)")
        report_lines.append("")

        report_lines.append("--- NOT FOUND (missing from sitemap) ---")
        if missing:
            report_lines.extend(f"✘ {u}" for u in missing)
        else:
            report_lines.append("(none)")
        report_lines.append("")
        report_lines.append("===== END OF REPORT =====")

        self.compare_output_box.setPlainText("\n".join(report_lines))

        # Log just the summary
        self.log(f"[COMPARE] Pasted: {len(pasted)} — Found: {len(found)} — Not found: {len(missing)}")

        # Save to file (Web Crawler style)
        self.save_compare_report_to_excel(pasted, found, missing)

        QMessageBox.information(
            self,
            "Compare Results",
            (
                f"Pasted (Excel): {len(pasted)}\n"
                f"Found in sitemap: {len(found)}\n"
                f"Not found: {len(missing)}\n\n"
                "Detailed output is shown below the Compare button and an Excel file was generated."
            )
        )

    def save_compare_report_to_excel(self, pasted, found, missing):
        # Ask where to save
        folder = QFileDialog.getExistingDirectory(self, "Select folder to save compare report")
        if not folder:
            self.log("[COMPARE] User cancelled saving Excel report.")
            return

        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"sitemap_compare_{ts}.xlsx"
        path = os.path.join(folder, filename)

        wb = Workbook()

        # Summary
        ws_sum = wb.active
        ws_sum.title = "Summary"
        ws_sum.append(["Metric", "Value"])
        ws_sum.append(["Total pasted (Excel)", len(pasted)])
        ws_sum.append(["Found in sitemap", len(found)])
        ws_sum.append(["Not found", len(missing)])

        # Found
        ws_found = wb.create_sheet("Found")
        ws_found.append(["URL"])
        for u in found:
            ws_found.append([u])

        # Not found
        ws_missing = wb.create_sheet("NotFound")
        ws_missing.append(["URL"])
        for u in missing:
            ws_missing.append([u])

        wb.save(path)
        self.log(f"[COMPARE] Excel report saved to: {path}")
