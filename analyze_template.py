"""
Script to analyze the metachecker_template.xlsx structure
"""

import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('metachecker_template.xlsx', data_only=True)
ws = wb.active

print(f"Sheet Name: {ws.title}")
print(f"Total Rows: {ws.max_row}")
print(f"Total Columns: {ws.max_column}")
print("=" * 80)

# Get headers (first row)
headers = []
for cell in ws[1]:
    headers.append(cell.value)

print("\nColumn Headers:")
for i, header in enumerate(headers, 1):
    print(f"  Column {i}: {header}")

print("\n" + "=" * 80)
print("\nFirst 3 data rows:")
for row_idx in range(2, min(5, ws.max_row + 1)):
    print(f"\nRow {row_idx}:")
    row_data = list(ws[row_idx])
    for i, (header, cell) in enumerate(zip(headers, row_data), 1):
        value = cell.value if cell.value is not None else "[EMPTY]"
        # Truncate long values
        if isinstance(value, str) and len(value) > 60:
            value = value[:60] + "..."
        print(f"  {header}: {value}")

print("\n" + "=" * 80)
print("\nColumn Name Analysis:")
for i, header in enumerate(headers, 1):
    if header:
        header_lower = str(header).lower()
        if 'url' in header_lower:
            print(f"  ✓ URL column found: {header} (column {i})")
        elif 'expected' in header_lower:
            print(f"  ✓ Expected column: {header} (column {i})")
        elif 'current' in header_lower:
            print(f"  ✓ Current column: {header} (column {i})")
