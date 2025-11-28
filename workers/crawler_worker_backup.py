"""
CrawlerThread worker for web crawling operations.
Fetches URLs, extracts specified data, and saves to Excel files.
"""

import os
import re
import datetime
import asyncio
import aiohttp
import openpyxl
from urllib.parse import urlparse
from bs4 import BeautifulSoup
from PyQt6.QtCore import QThread, pyqtSignal

from config import MAX_EXCEL_CELL_LENGTH, TIMEOUT_STANDARD


class CrawlerThread(QThread):
    """
    A worker thread for crawling websites asynchronously.
    Fetches URLs, extracts specified data, and saves it to Excel files.
    """
    progress_update = pyqtSignal(int)
    log_update = pyqtSignal(str)
    finished = pyqtSignal(str)

    def __init__(self, mode, search_input, urls, extract_options, check_errors, output_folder):
        super().__init__()
        self.mode = mode  # 1: Search modules (classes), 2: Search words, 0: No search
        self.search_input = search_input
        self.urls = urls
        self.extract_options = extract_options
        self.check_errors = check_errors
        self.output_folder = output_folder or f"web_crawler_results_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
        os.makedirs(self.output_folder, exist_ok=True)
        self.total_pages_crawled = 0
        self.stopped = False

    def stop(self):
        """Signals the thread to stop processing."""
        self.log_update.emit("Stopping crawler...")
        self.stopped = True

    def run(self):
        """Runs the asynchronous crawling process."""
        try:
            asyncio.run(self.main())
        except asyncio.CancelledError:
            self.log_update.emit("Crawling was cancelled")
        except Exception as e:
            self.log_update.emit(f"An unexpected error occurred: {e}")

    async def main(self):
        """Main async function to set up and execute crawling tasks."""
        class_patterns = []
        search_patterns = []

        if self.mode == 1 and self.search_input:
            class_patterns = [pattern.strip() for pattern in self.search_input.split(',') if pattern.strip()]
        elif self.mode == 2 and self.search_input:
            search_patterns = self._generate_search_patterns([word.strip() for word in self.search_input.split(',') if word.strip()])

        main_filename = os.path.join(self.output_folder, "results.xlsx")
        error_filename = os.path.join(self.output_folder, "error_results.xlsx") if self.check_errors else None
        content_filename = os.path.join(self.output_folder, "content_results.xlsx") if self.extract_options["content"] else None

        wb_main = openpyxl.Workbook()
        ws_main = wb_main.active
        ws_main.title = "Main Results"
        headers = ["URL"]
        if self.extract_options["title"]: headers.append("Title")
        if self.extract_options["meta_title"]: headers.append("Meta Title")
        if self.extract_options["meta_description"]: headers.append("Meta Description")
        if self.extract_options["content"]: headers.append("Content Snippet")
        if self.extract_options["meta_tags"]: headers.append("Meta Tags")
        if self.mode == 1: headers.append("Module Found")
        elif self.mode == 2: headers.append("Found Words")
        ws_main.append(headers)

        wb_content = openpyxl.Workbook() if content_filename else None
        ws_content = wb_content.active if wb_content else None
        if ws_content:
            ws_content.title = "Content Results"
            ws_content.append(["URL", "Full Content"])

        wb_errors = openpyxl.Workbook() if error_filename else None
        ws_errors = wb_errors.active if wb_errors else None
        if ws_errors:
            ws_errors.title = "Error Results"
            ws_errors.append(["URL", "Status Code", "Redirect"])

        async with aiohttp.ClientSession() as session:
            all_urls_to_check = []
            for url in self.urls:
                if self.stopped: break
                if urlparse(url).path.endswith(".xml"):
                    self.log_update.emit(f"Fetching URLs from sitemap: {url}")
                    sitemap_urls = await self._get_sitemap_urls(url, session)
                    all_urls_to_check.extend(sitemap_urls)
                else:
                    all_urls_to_check.append(url)
            
            total_urls = len(all_urls_to_check)
            self.total_pages_crawled = total_urls
            
            tasks = [self._crawl_url(u, session, class_patterns, search_patterns) for u in all_urls_to_check]
            
            for i, future in enumerate(asyncio.as_completed(tasks)):
                if  self.stopped: break
                
                result = await future
                if result:
                    if result['type'] == 'success':
                        ws_main.append(result['main_data'])
                        if ws_content and 'content_data' in result:
                            ws_content.append(result['content_data'])
                    elif result['type'] == 'error' and ws_errors:
                        ws_errors.append(result['error_data'])

                progress = int((i + 1) / total_urls * 100) if total_urls > 0 else 100
                self.progress_update.emit(progress)
                self.log_update.emit(f"Processed {i + 1}/{total_urls} URLs")

        if not self.stopped:
            self.log_update.emit("Saving results to Excel files...")
            wb_main.save(main_filename)
            if wb_content: wb_content.save(content_filename)
            if wb_errors: wb_errors.save(error_filename)
            
            self.log_update.emit(f"Crawling completed. Results saved to {self.output_folder}")
            self.log_update.emit(f"Total pages processed: {self.total_pages_crawled}")
            self.finished.emit(self.output_folder)
        else:
            self.log_update.emit("Crawling stopped by user.")

    def _generate_search_patterns(self, words):
        return [re.compile(re.escape(word), re.IGNORECASE) for word in words]

    async def _get_sitemap_urls(self, sitemap_url, session):
        try:
            async with session.get(sitemap_url, ssl=False) as response:
                if response.status == 200:
                    content = await response.text()
                    soup = BeautifulSoup(content, 'lxml-xml')
                    return [loc.text for loc in soup.find_all('loc')]
                else:
                    self.log_update.emit(f"Sitemap fetch failed for {sitemap_url}: Status {response.status}")
                    return []
        except aiohttp.ClientError as e:
            self.log_update.emit(f"Network error reading sitemap {sitemap_url}: {e}")
            return []
        except Exception as e:
            self.log_update.emit(f"Error reading sitemap {sitemap_url}: {e}")
            return []

    async def _crawl_url(self, url, session, class_patterns, search_patterns):
        try:
            async with session.get(url, ssl=False, timeout=TIMEOUT_STANDARD) as response:
                if response.status in {403, 404} and self.check_errors:
                    redirect_url = response.headers.get("Location", "N/A")
                    self.log_update.emit(f"Error {response.status} for {url}")
                    return {'type': 'error', 'error_data': [url, response.status, redirect_url]}

                if response.status == 200:
                    html = await response.text()
                    soup = BeautifulSoup(html, 'lxml')
                    result = {'type': 'success'}
                    row_data = [url]
                    
                    if self.extract_options["title"]:
                        row_data.append(soup.title.string.strip() if soup.title and soup.title.string else "No title")
                    if self.extract_options["meta_title"]:
                        meta = soup.find("meta", property="og:title")
                        row_data.append(meta["content"] if meta and meta.get("content") else "No meta title")
                    if self.extract_options["meta_description"]:
                        meta = soup.find("meta", property="og:description")
                        row_data.append(meta["content"] if meta and meta.get("content") else "No meta description")
                    
                    if self.extract_options["content"]:
                        content = soup.get_text(separator=' ', strip=True)
                        snippet = content[:MAX_EXCEL_CELL_LENGTH]
                        row_data.append(snippet)
                        result['content_data'] = [url, content]

                    if self.extract_options["meta_tags"]:
                        tags = [f"{meta.get('name') or meta.get('property')}: {meta.get('content', '')}" 
                                for meta in soup.find_all("meta") if meta.get("name") or meta.get("property")]
                        row_data.append(", ".join(tags))
                        
                    if self.mode == 1:
                        found = any(soup.find("div", class_=cp) for cp in class_patterns)
                        row_data.append("Yes" if found else "No")
                    elif self.mode == 2:
                        text = soup.get_text()
                        found_words = [p.pattern for p in search_patterns if p.search(text)]
                        row_data.append(', '.join(found_words) if found_words else "None")

                    result['main_data'] = row_data
                    return result
                else:
                    self.log_update.emit(f"Non-200 status for {url}: {response.status}")

        except asyncio.TimeoutError:
            self.log_update.emit(f"Timeout processing {url}")
        except aiohttp.ClientError as e:
            self.log_update.emit(f"Network error for {url}: {e}")
        except Exception as e:
            self.log_update.emit(f"Failed to process {url}: {e}")
        return None
