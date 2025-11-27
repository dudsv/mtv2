"""
Test script to identify issues with MetaCheckWorker's Excel reading logic
"""

import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('metachecker_template.xlsx', data_only=True, read_only=True)

# Try to find the right sheet
ws = wb.active
print(f"Active sheet: {ws.title}")
print(f"All sheets: {wb.sheetnames}")
print()

# Look for sheets that might contain our data
for name in wb.sheetnames:
    low = name.lower()
    if "h1" in low or "update" in low or "meta" in low:
        print(f"Found relevant sheet: {name}")
        ws = wb[name]
        break

# Get all rows
rows = list(ws.iter_rows(values_only=True))
print(f"Total rows: {len(rows)}")
print()

# Function to normalize cell values for searching
def norm(cell):
    return str(cell).strip().lower() if cell is not None else ""

# Find header row
header_row_idx = None
max_scan_header_rows = min(20, len(rows))
for i in range(max_scan_header_rows):
    row = rows[i]
    for cell in row:
        h = norm(cell)
        if "page url" in h or h == "url":
            header_row_idx = i
            break
    if header_row_idx is not None:
        break

if header_row_idx is None:
    print("ERROR: No header row found!")
    exit(1)

print(f"Found header at row: {header_row_idx + 1}")
header = rows[header_row_idx]

# Print all headers
print("\nAll column headers:")
for idx, h in enumerate(header):
    print(f"  Column {idx}: {h}")

# Detect column indices
url_idx = h1_idx = mt_idx = md_idx = ogt_idx = ogd_idx = None

for idx, cell in enumerate(header):
    h = norm(cell)
    
    # URL
    if url_idx is None and ("page url" in h or h == "url"):
        url_idx = idx
        print(f"\n✓ URL column: {idx} - '{header[idx]}'")
    
    # Expected H1
    if h1_idx is None:
        if "expected" in h and "h1" in h:
            h1_idx = idx
            print(f"✓ Expected H1 column: {idx} - '{header[idx]}'")
        elif "expected" in h and "heading 1" in h:
            h1_idx = idx
            print(f"✓ Expected H1 column: {idx} - '{header[idx]}'")
    
    # Expected Meta Title
    if mt_idx is None:
        if "expected" in h and "meta" in h and "title" in h:
            mt_idx = idx
            print(f"✓ Expected Meta Title column: {idx} - '{header[idx]}'")
        elif "expected" in h and "title" in h and "page" in h:
            mt_idx = idx
            print(f"✓ Expected Meta Title column: {idx} - '{header[idx]}'")
    
    # Expected Meta Description
    if md_idx is None:
        if "expected" in h and "description" in h and "og" not in h:
            md_idx = idx
            print(f"✓ Expected Meta Description column: {idx} - '{header[idx]}'")
    
    # Expected OG Title
    if ogt_idx is None:
        if "expected" in h and "og" in h and "title" in h:
            ogt_idx = idx
            print(f"✓ Expected OG Title column: {idx} - '{header[idx]}'")
    
    # Expected OG Description
    if ogd_idx is None:
        if "expected" in h and "og" in h and "description" in h:
            ogd_idx = idx
            print(f"✓ Expected OG Description column: {idx} - '{header[idx]}'")

print("\n" + "="*80)
print("Column Detection Summary:")
print(f"URL_IDX={url_idx}")
print(f"META_TITLE_IDX={mt_idx}")
print(f"META_DESC_IDX={md_idx}")  
print(f"OG_TITLE_IDX={ogt_idx}")
print(f"OG_DESC_IDX={ogd_idx}")
print(f"H1_IDX={h1_idx}")
print("="*80)

wb.close()
